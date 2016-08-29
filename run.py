# python scholar.py -c 1 -p "The obstinate audience" -t -a "Bauer" --after 1963 --before 1965
## python scholar.py -c 1 -p -t
# import openpyxl
#
# wb = openpyxl.load_workbook('sample.xlsx')
# sheets = wb.get_sheet_names()
# sheet = wb.get_sheet_by_name(sheets[0])
# wbw = openpyxl.Workbook()
# ws = wbw.active
#
# row = 2
# while True:
#     title = sheet['B' + str(row)].value
#     journal = sheet['C' + str(row)].value
#     author = sheet['E' + str(row)].value
#     year = sheet['N' + str(row)].value
#     if title is None:
#         break
#     row += 1
#     ws.append([title, journal, author, year])
#
# wbw.save("output.xlsx")

import sqlite3
import openpyxl
from scholar import get_database_pubs
import os

global conn
conn = None
db_path = 'pubs_db3'
filename = 'pubs_miss.xlsx'
filename2 = 'pubs_match.xlsx'
try:
    try:
        os.remove(filename)
        os.remove(filename2)
        print("File Removed!")
    except:
        pass
    get_database_pubs(db_path)
    conn = sqlite3.connect(db_path)
except Exception, e:
    print(repr(e))
    exit(0)
wbw = openpyxl.Workbook()
ws = wbw.active
wbw2 = openpyxl.Workbook()
ws2 = wbw2.active
wb = openpyxl.load_workbook('E:\crawling\sample.xlsx')
sheets = wb.get_sheet_names()
sheet = wb.get_sheet_by_name(sheets[0])
row = 2
try:
    while True:
        print("############  " + str(row) + "  ###################")
        title = sheet['B' + str(row)].value
        journal = sheet['C' + str(row)].value
        author = sheet['E' + str(row)].value
        year = sheet['N' + str(row)].value
        if title is None:
            break
        print(title)
        cur = conn.cursor()
        cur.execute('SELECT id, cluster_id, title, authors, year FROM pub WHERE title LIKE ?', (title.strip(),))
        pb = cur.fetchone()
        # for pub_id in data:
        #     print(pub_id[0])
        #     cur.execute('SELECT cluster_id, title FROM pub WHERE id=?', pub_id)
    #     pb = cur.fetchone()
        row += 1
        if pb is None or len(pb) == 0:
            ws.append([title, journal, author, year])
            continue
        ws2.append([pb[0], pb[1], unicode(pb[2]), unicode(pb[3]), pb[4]])
finally:
    wbw.save(filename)
    wbw2.save(filename2)
    if conn:
        conn.close()
